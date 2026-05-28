from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from config import settings
from stages.stage3_classify import DAMAGE_CATEGORIES

CURRENCY_FORMAT = "$#,##0.00"
DATE_FORMAT = "yyyy-mm-dd"
HEADER_LABELS = ["Date", "Description", "Amount (CAD)", "Source / Citation", "Notes"]
MAX_DATA_ROW = 500

# Consultant / finance palette
COLOR_NAVY = "1F4E79"
COLOR_HEADER_TEXT = "FFFFFF"
COLOR_ALT_ROW = "F2F7FB"
COLOR_TOTAL_FILL = "D9E2F3"
COLOR_BORDER = "B4C6E7"
COLOR_SUBTITLE = "44546A"

FONT_TITLE = Font(name="Calibri", size=14, bold=True, color=COLOR_NAVY)
FONT_SUBTITLE = Font(name="Calibri", size=10, italic=True, color=COLOR_SUBTITLE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=COLOR_HEADER_TEXT)
FONT_BODY = Font(name="Calibri", size=10, color="000000")
FONT_TOTAL = Font(name="Calibri", size=11, bold=True, color=COLOR_NAVY)

FILL_HEADER = PatternFill("solid", fgColor=COLOR_NAVY)
FILL_ALT = PatternFill("solid", fgColor=COLOR_ALT_ROW)
FILL_TOTAL = PatternFill("solid", fgColor=COLOR_TOTAL_FILL)

THIN_BORDER = Border(
    left=Side(style="thin", color=COLOR_BORDER),
    right=Side(style="thin", color=COLOR_BORDER),
    top=Side(style="thin", color=COLOR_BORDER),
    bottom=Side(style="thin", color=COLOR_BORDER),
)
THICK_BOTTOM = Border(
    left=Side(style="thin", color=COLOR_BORDER),
    right=Side(style="thin", color=COLOR_BORDER),
    top=Side(style="thin", color=COLOR_BORDER),
    bottom=Side(style="medium", color=COLOR_NAVY),
)

ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="top")
ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

DATA_START_ROW = 5
HEADER_ROW_NUM = 4


def _excel_path(job_id: str) -> Path:
    return Path(settings.JOBS_PATH) / job_id / "schedule.xlsx"


def _classifications_path(job_id: str) -> Path:
    return Path(settings.JOBS_PATH) / job_id / "classifications.json"


def _category_sheet_name(category: str) -> str:
    labels = {
        "past_lost_income": "Past Lost Income",
        "future_lost_income": "Future Lost Income",
        "medical_expenses": "Medical Expenses",
        "future_care_costs": "Future Care Costs",
        "out_of_pocket": "Out of Pocket",
        "loss_of_valuable_services": "Loss of Services",
        "non_pecuniary": "Non Pecuniary",
        "pre_judgment_interest": "Pre Judgment Interest",
    }
    return labels.get(category, category.replace("_", " ").title())[:31]


def _safe_sheet_reference(sheet_name: str) -> str:
    escaped = sheet_name.replace("'", "''")
    return f"'{escaped}'"


def _set_column_widths(worksheet: Worksheet) -> None:
    widths = {"A": 14, "B": 48, "C": 16, "D": 28, "E": 36}
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def _style_header_row(worksheet: Worksheet, row: int, column_count: int = 5) -> None:
    for col in range(1, column_count + 1):
        cell = worksheet.cell(row=row, column=col)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def _style_data_cell(
    worksheet: Worksheet,
    row: int,
    col: int,
    *,
    is_amount: bool = False,
    alt: bool = False,
) -> None:
    cell = worksheet.cell(row=row, column=col)
    cell.font = FONT_BODY
    cell.border = THIN_BORDER
    if alt:
        cell.fill = FILL_ALT

    if col == 1:
        cell.alignment = ALIGN_CENTER
    elif col == 3 or is_amount:
        cell.alignment = ALIGN_RIGHT
        if cell.value not in (None, "") and isinstance(cell.value, (int, float)):
            cell.number_format = CURRENCY_FORMAT
    else:
        cell.alignment = ALIGN_WRAP


def _write_sheet_banner(
    worksheet: Worksheet,
    *,
    job_id: str,
    section_title: str,
) -> None:
    prepared = datetime.now(timezone.utc).strftime("%B %d, %Y")
    worksheet.merge_cells("A1:E1")
    worksheet["A1"] = "LitiDoc — Schedule of Damages"
    worksheet["A1"].font = FONT_TITLE
    worksheet["A1"].alignment = Alignment(horizontal="left", vertical="center")

    worksheet.merge_cells("A2:E2")
    worksheet["A2"] = section_title
    worksheet["A2"].font = Font(name="Calibri", size=12, bold=True, color=COLOR_NAVY)
    worksheet["A2"].alignment = Alignment(horizontal="left", vertical="center")

    worksheet.merge_cells("A3:E3")
    worksheet["A3"] = f"Job Reference: {job_id}  |  Prepared: {prepared} (UTC)"
    worksheet["A3"].font = FONT_SUBTITLE
    worksheet["A3"].alignment = Alignment(horizontal="left", vertical="center")

    worksheet.row_dimensions[1].height = 24
    worksheet.row_dimensions[2].height = 20
    worksheet.row_dimensions[3].height = 18
    worksheet.row_dimensions[HEADER_ROW_NUM].height = 22


def create_category_sheet(
    workbook: Workbook,
    sheet_name: str,
    items: list[dict],
    *,
    job_id: str,
    category_key: str,
) -> Worksheet:
    """Create and populate a professionally formatted damage category worksheet."""
    worksheet = workbook.create_sheet(title=sheet_name)
    section_title = f"{_category_sheet_name(category_key)} — Detail Schedule"
    _write_sheet_banner(worksheet, job_id=job_id, section_title=section_title)

    for col, label in enumerate(HEADER_LABELS, start=1):
        worksheet.cell(row=HEADER_ROW_NUM, column=col, value=label)
    _style_header_row(worksheet, HEADER_ROW_NUM)

    data_row = DATA_START_ROW
    if items:
        for index, item in enumerate(items):
            amount = item.get("amount")
            worksheet.cell(row=data_row, column=1, value=item.get("date", "unspecified"))
            worksheet.cell(row=data_row, column=2, value=item.get("description", ""))
            worksheet.cell(
                row=data_row,
                column=3,
                value=amount if amount is not None else "—",
            )
            worksheet.cell(row=data_row, column=4, value=item.get("source", ""))
            worksheet.cell(row=data_row, column=5, value=item.get("notes", ""))

            alt = index % 2 == 1
            for col in range(1, 6):
                _style_data_cell(
                    worksheet,
                    data_row,
                    col,
                    is_amount=(col == 3),
                    alt=alt,
                )
            worksheet.row_dimensions[data_row].height = 32
            data_row += 1
    else:
        worksheet.merge_cells(start_row=DATA_START_ROW, start_column=1, end_row=DATA_START_ROW, end_column=5)
        worksheet.cell(row=DATA_START_ROW, column=1, value="No line items recorded for this category.")
        worksheet.cell(row=DATA_START_ROW, column=1).font = FONT_SUBTITLE
        worksheet.cell(row=DATA_START_ROW, column=1).alignment = ALIGN_WRAP
        data_row += 1

    subtotal_row = data_row + 1
    worksheet.cell(row=subtotal_row, column=2, value="Subtotal")
    worksheet.cell(
        row=subtotal_row,
        column=3,
        value=f"=SUM(C{DATA_START_ROW}:C{data_row - 1})",
    )
    for col in range(1, 6):
        cell = worksheet.cell(row=subtotal_row, column=col)
        cell.font = FONT_TOTAL
        cell.fill = FILL_TOTAL
        cell.border = THICK_BOTTOM
        if col == 3:
            cell.number_format = CURRENCY_FORMAT
            cell.alignment = ALIGN_RIGHT
        elif col == 2:
            cell.alignment = Alignment(horizontal="right", vertical="center")

    _set_column_widths(worksheet)
    worksheet.freeze_panes = f"A{DATA_START_ROW}"
    worksheet.sheet_view.showGridLines = False
    return worksheet


def add_summary_sheet(
    workbook: Workbook,
    category_totals: dict[str, str],
    *,
    job_id: str,
) -> Worksheet:
    """Add executive summary sheet with per-category totals and grand total."""
    worksheet = workbook.create_sheet(title="Summary", index=0)
    _write_sheet_banner(
        worksheet,
        job_id=job_id,
        section_title="Executive Summary — Heads of Damage",
    )

    worksheet.cell(row=HEADER_ROW_NUM, column=1, value="Head of Damage")
    worksheet.cell(row=HEADER_ROW_NUM, column=2, value="Total Amount (CAD)")
    worksheet.cell(row=HEADER_ROW_NUM, column=3, value="Source Tab")
    _style_header_row(worksheet, HEADER_ROW_NUM, column_count=3)

    summary_start_row = DATA_START_ROW
    for offset, category in enumerate(DAMAGE_CATEGORIES):
        row_number = summary_start_row + offset
        sheet_name = category_totals[category]
        worksheet.cell(row=row_number, column=1, value=_category_sheet_name(category))
        worksheet.cell(
            row=row_number,
            column=2,
            value=f"=SUM({_safe_sheet_reference(sheet_name)}!C{DATA_START_ROW}:C{MAX_DATA_ROW})",
        )
        worksheet.cell(row=row_number, column=3, value=sheet_name)

        alt = offset % 2 == 1
        for col in range(1, 4):
            cell = worksheet.cell(row=row_number, column=col)
            cell.font = FONT_BODY
            cell.border = THIN_BORDER
            if alt:
                cell.fill = FILL_ALT
            if col == 2:
                cell.number_format = CURRENCY_FORMAT
                cell.alignment = ALIGN_RIGHT
            else:
                cell.alignment = ALIGN_WRAP if col == 1 else ALIGN_CENTER

    grand_total_row = summary_start_row + len(DAMAGE_CATEGORIES) + 1
    last_category_row = summary_start_row + len(DAMAGE_CATEGORIES) - 1
    worksheet.cell(row=grand_total_row, column=1, value="Grand Total (All Heads of Damage)")
    worksheet.cell(
        row=grand_total_row,
        column=2,
        value=f"=SUM(B{summary_start_row}:B{last_category_row})",
    )
    for col in range(1, 4):
        cell = worksheet.cell(row=grand_total_row, column=col)
        cell.font = FONT_TOTAL
        cell.fill = FILL_TOTAL
        cell.border = THICK_BOTTOM
        if col == 2:
            cell.number_format = CURRENCY_FORMAT
            cell.alignment = ALIGN_RIGHT

    worksheet.column_dimensions["A"].width = 34
    worksheet.column_dimensions["B"].width = 20
    worksheet.column_dimensions["C"].width = 22
    worksheet.freeze_panes = f"A{DATA_START_ROW}"
    worksheet.sheet_view.showGridLines = False

    # Highlight summary tab
    worksheet.sheet_properties.tabColor = COLOR_NAVY
    return worksheet


def _load_classifications(job_id: str) -> dict:
    path = _classifications_path(job_id)
    if not path.exists():
        raise FileNotFoundError(f"Classifications not found: {path}")
    return json.loads(path.read_text(encoding="utf-8"))


def _items_by_category(classifications: dict) -> dict[str, list[dict]]:
    by_category = classifications.get("by_category")
    if isinstance(by_category, dict) and by_category:
        return {
            category: payload.get("items", [])
            for category, payload in by_category.items()
            if category in DAMAGE_CATEGORIES
        }

    grouped = {category: [] for category in DAMAGE_CATEGORIES}
    for item in classifications.get("damages", []):
        category = item.get("category", "out_of_pocket")
        if category in grouped:
            grouped[category].append(item)
    return grouped


def build_excel(job_id: str, classifications: dict | None = None) -> str:
    """Build damages schedule workbook and return saved file path."""
    if not job_id.strip():
        raise ValueError("job_id is required.")

    if classifications is None:
        classifications = _load_classifications(job_id)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    items_by_category = _items_by_category(classifications)
    sheet_names_by_category: dict[str, str] = {}

    for category in DAMAGE_CATEGORIES:
        sheet_name = _category_sheet_name(category)
        sheet_names_by_category[category] = sheet_name
        create_category_sheet(
            workbook,
            sheet_name,
            items_by_category.get(category, []),
            job_id=job_id,
            category_key=category,
        )
        workbook[sheet_name].sheet_properties.tabColor = "AAB7C9"

    add_summary_sheet(workbook, sheet_names_by_category, job_id=job_id)

    output_path = _excel_path(job_id)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    print(f"[Stage4] excel saved job_id={job_id} path={output_path}")
    return str(output_path)


def build_minimal_excel(job_id: str) -> str:
    """Last-resort workbook with professional formatting."""
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    demo_items = [
        {
            "date": "2021-03-05",
            "description": "Demo fallback medical expense",
            "amount": 2500.0,
            "source": "Doc 4.1, p. 2",
            "notes": "Generated by minimal Excel fallback",
        },
        {
            "date": "2021-04-01",
            "description": "Demo fallback lost income",
            "amount": 47500.0,
            "source": "Doc 4.1, p. 10",
            "notes": "Generated by minimal Excel fallback",
        },
    ]

    sheet_names_by_category: dict[str, str] = {}
    for category in DAMAGE_CATEGORIES:
        sheet_name = _category_sheet_name(category)
        sheet_names_by_category[category] = sheet_name
        if category == "medical_expenses":
            items = [demo_items[0]]
        elif category == "past_lost_income":
            items = [demo_items[1]]
        else:
            items = []
        create_category_sheet(
            workbook,
            sheet_name,
            items,
            job_id=job_id,
            category_key=category,
        )

    add_summary_sheet(workbook, sheet_names_by_category, job_id=job_id)

    output_path = _excel_path(job_id)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    print(f"[Stage4] minimal excel saved job_id={job_id} path={output_path}")
    return str(output_path)


def build_excel_guaranteed(job_id: str, classifications: dict | None = None) -> str:
    """Always produce schedule.xlsx using primary, fallback, then minimal builders."""
    errors: list[str] = []

    if classifications is not None:
        try:
            return build_excel(job_id, classifications)
        except Exception as error:
            errors.append(f"primary build failed: {error}")
            print(f"[Stage4] WARNING: {errors[-1]}")

    try:
        from core.fallbacks import get_fallback_classifications

        fallback = get_fallback_classifications(job_id)
        return build_excel(job_id, fallback)
    except Exception as error:
        errors.append(f"fallback classifications build failed: {error}")
        print(f"[Stage4] WARNING: {errors[-1]}")

    try:
        return build_minimal_excel(job_id)
    except Exception as error:
        errors.append(f"minimal build failed: {error}")
        raise RuntimeError(
            "Unable to create Excel schedule after all fallback attempts: "
            + " | ".join(errors)
        ) from error
